"""Backfill teneur en eau — import du relevé terrain historique
(data_teneur/Teneur en eau Paroi.xlsx) dans InfluxDB.

Réutilise le mapping mur/couche déjà validé et documenté (logique_projet.md
section 16, extrait à l'identique de data_reel_compile/extraire_teneur_eau_reel.py)
— mais écrit directement dans mesures_teneur_eau au lieu d'un JSON
intermédiaire POC, avec les tags de production (nom_mur "SOCMA 1"/"SOCMA 2",
pas la numérotation interne 1/2 du POC).

64 mesures réelles, 11 relevés terrain (Protimeter/humidimètre à pointes) du
21/11/2025 au 18/03/2026, 2 murs x 3 couches (carreau_ext, carreau_isolant,
milieu_isolant — milieu_carreau et isolant_osb n'ont aucune mesure réelle,
pas de point généré pour elles).

Identité de saisie : ces mesures n'ont pas de session utilisateur réelle
(relevé terrain historique, pas une saisie via l'appli) — utilisateur_id/
utilisateur_nom fixés à une valeur explicite "backfill_historique", jamais
confondue avec une vraie saisie future (cf. logique_projet.md section 16).

Écriture directe InfluxDB (line protocol), sans MQTT/Kafka — déjà tranché
dans la conception (section 16) : une saisie humaine ponctuelle n'a pas
besoin de la chaîne de résilience prévue pour un flux continu.

Usage :
    python backfill_teneur_eau.py                  Aperçu (dry-run) — n'écrit rien
    python backfill_teneur_eau.py --confirmer       Écriture réelle dans InfluxDB

    Variables d'environnement (mêmes défauts que backfill_hr_t.py) :
        INFLUX_URL, INFLUX_TOKEN, INFLUX_ORG, INFLUX_BUCKET
        TENEUR_EAU_SOURCE  Chemin du classeur (défaut : data_teneur/Teneur en eau Paroi.xlsx à côté de ce script)
"""
import os
import sys
from datetime import datetime, timezone

import openpyxl
from influxdb_client import InfluxDBClient, WritePrecision
from influxdb_client.client.write_api import SYNCHRONOUS

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Configuration.
# ---------------------------------------------------------------------------
INFLUX_URL = os.getenv("INFLUX_URL", "http://localhost:8086")
INFLUX_TOKEN = os.getenv("INFLUX_TOKEN", "MON_TOKEN_API_GENERE_PAR_INFLUXDB")
INFLUX_ORG = os.getenv("INFLUX_ORG", "FRD_CODEM")
INFLUX_BUCKET = os.getenv("INFLUX_BUCKET", "Capteurs")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TENEUR_EAU_SOURCE = os.getenv(
    "TENEUR_EAU_SOURCE", os.path.join(SCRIPT_DIR, "data_teneur", "Teneur en eau Paroi.xlsx")
)

CONFIRMER = "--confirmer" in sys.argv

MESURE_TENEUR_EAU = "mesures_teneur_eau"

UTILISATEUR_ID = "backfill_historique"
UTILISATEUR_NOM = "Import historique (data_teneur, 12/08/2026)"
COMMENTAIRE = "Relevé terrain historique (Teneur en eau Paroi.xlsx), importé rétroactivement"

# Mapping mur/couche — identique à extraire_teneur_eau_reel.py, mais le mur
# pointe directement vers le nom de production (SOCMA 1/SOCMA 2) plutôt que
# la numérotation interne 1/2 du POC.
COUCHE_PAR_PREFIXE = {
    "carreau ext": "carreau_ext",
    "carreau int": "carreau_isolant",
    "isolant": "milieu_isolant",
}
MUR_PAR_SUFFIXE = {
    "normal": "SOCMA 1",
    "gros gravier": "SOCMA 2",
}


def _echap_tag(valeur: str) -> str:
    return valeur.replace("\\", "\\\\").replace(",", "\\,").replace("=", "\\=").replace(" ", "\\ ")


def _echap_field_str(valeur: str) -> str:
    return valeur.replace("\\", "\\\\").replace('"', '\\"')


def parse_label(label: str):
    """Reconnaît un libellé de ligne source (ex. 'Carreau ext normal') en
    (mur, couche) — même logique que extraire_teneur_eau_reel.py."""
    label = label.strip().lower()
    for suffixe, mur in MUR_PAR_SUFFIXE.items():
        if label.endswith(suffixe):
            prefixe = label[: -len(suffixe)].strip()
            if prefixe in COUCHE_PAR_PREFIXE:
                return mur, COUCHE_PAR_PREFIXE[prefixe]
    return None, None


def lire_mesures():
    """Génère (date, mur, couche, teneur_eau_pourcent) pour chaque mesure
    valide du classeur source."""
    wb = openpyxl.load_workbook(TENEUR_EAU_SOURCE, data_only=True)
    ws = wb["Feuil1"]

    # Ligne des dates : la première contenant au moins 2 cellules datetime.
    ligne_dates, colonnes = None, None
    for r in range(1, ws.max_row + 1):
        cols_dates = [
            c for c in range(1, ws.max_column + 1)
            if hasattr(ws.cell(row=r, column=c).value, "year")
        ]
        if len(cols_dates) >= 2:
            ligne_dates, colonnes = r, cols_dates
            break
    if ligne_dates is None:
        raise RuntimeError("Ligne d'en-tête des dates introuvable dans Feuil1.")

    dates = {
        c: ws.cell(row=ligne_dates, column=c).value.replace(tzinfo=timezone.utc)
        for c in colonnes
    }

    non_reconnues = []
    for r in range(ligne_dates + 1, ws.max_row + 1):
        label = None
        for c in range(1, min(colonnes)):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip():
                label = v
                break
        if label is None:
            continue
        mur, couche = parse_label(label)
        if mur is None:
            non_reconnues.append(label)
            continue
        for c in colonnes:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            yield dates[c], mur, couche, round(v * 100, 3)

    if non_reconnues:
        print(f"  ⚠️  Lignes ignorées (libellé non reconnu) : {non_reconnues}")


def construire_ligne(d: datetime, mur: str, couche: str, valeur: float) -> str:
    tags = (
        f"utilisateur_id={_echap_tag(UTILISATEUR_ID)},"
        f"utilisateur_nom={_echap_tag(UTILISATEUR_NOM)},"
        f"mur={_echap_tag(mur)},"
        f"couche={_echap_tag(couche)},"
        f"prestation={_echap_tag('Non défini')}"
    )
    fields = (
        f"teneur_eau_pourcent={valeur},"
        f'commentaire="{_echap_field_str(COMMENTAIRE)}"'
    )
    ts_ns = int(d.timestamp() * 1_000_000_000)
    return f"{MESURE_TENEUR_EAU},{tags} {fields} {ts_ns}"


def main() -> None:
    print(f"Source : {TENEUR_EAU_SOURCE}")
    mesures = list(lire_mesures())

    lignes = [construire_ligne(d, mur, couche, valeur) for d, mur, couche, valeur in mesures]

    par_mur_couche = {}
    for d, mur, couche, valeur in mesures:
        par_mur_couche.setdefault((mur, couche), []).append((d.date().isoformat(), valeur))

    print(f"\n{len(lignes)} mesures trouvées, {len(par_mur_couche)} série(s) mur/couche :\n")
    for (mur, couche), points in sorted(par_mur_couche.items()):
        dates_str = ", ".join(f"{dt}={v}%" for dt, v in points)
        print(f"  {mur} / {couche} ({len(points)} points) : {dates_str}")

    print(f"\nTotal : {len(lignes)} points à écrire dans '{MESURE_TENEUR_EAU}' (bucket {INFLUX_BUCKET}).")

    if not CONFIRMER:
        print("\nAperçu uniquement (dry-run) — rien n'a été écrit dans InfluxDB.")
        print("Relancer avec --confirmer pour une écriture réelle.")
        if lignes:
            print("\nExemple de ligne générée :")
            print(" ", lignes[0])
        return

    print("\nÉcriture dans InfluxDB (mode synchrone)...")
    client = InfluxDBClient(url=INFLUX_URL, token=INFLUX_TOKEN, org=INFLUX_ORG)
    write_api = client.write_api(write_options=SYNCHRONOUS)
    try:
        write_api.write(
            bucket=INFLUX_BUCKET, org=INFLUX_ORG, record=lignes,
            write_precision=WritePrecision.NS,
        )
    finally:
        write_api.close()
        client.close()
    print(f"{len(lignes)} points écrits.")


if __name__ == "__main__":
    main()
